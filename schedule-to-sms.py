import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from twilio.rest import Client

# Sökväg till filen
filepath = 'updated_schedule.xlsx'

# Läs in Excel-filen
if os.path.exists(filepath):
    print(f"Filen {filepath} finns.")
else:
    print(f"Filen {filepath} saknas.")
    exit()

updated_schedule = filepath

# Öppna filen och använd det aktiva arket
wb1 = load_workbook(updated_schedule)
ws1 = wb1.active

# Spara personerna
person_data = []

# Iterera genom cellerna i första arket och hämta namn, mail och telefonnummer
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=2, max_col=ws1.max_column):
    for cell in row:
        # Kollar om cellen har text
        if cell.value is not None:
            # Dela upp cellvärdet baserat på newline
            values = cell.value.split('\n')
            # Säkerställ att det finns tillräckligt många värden
            if len(values) >= 3:  
                # Lägg till rum och tid
                person = {
                    "email": values[0],
                    "namn": values[1],
                    "telefon": values[2],
                    "rum": ws1.cell(row=1, column=cell.column).value,
                    "tid": ws1.cell(row=cell.row, column=1).value
                }
                person_data.append(person)

# Skriv ut för att kontrollera att data har extraherats korrekt

# Kontrollera att allt är med
for person in person_data:
    print(person)

### TODO ###
# Implementera Twilio
# SMS ska skickas till korrekt nummer med rum och tid

# Find your Account SID and Auth Token at twilio.com/console
# and set the environment variables. See http://twil.io/secure
#account_sid = os.environ["TWILIO_ACCOUNT_SID"]
#auth_token = os.environ["TWILIO_AUTH_TOKEN"]
#client = Client(account_sid, auth_token)

#from_number = "+4601234567"

# Skicka sms till varje person
for person in person_data:
    text = f"Hej {person['namn']}! Du är välkommen till oss {person['tid']} i {person['rum']}."
    print(text)
#    message = client.messages.create(
#        body=text,
#        from_=from_number,
#        to=person["telefon"],
#    )
#
#print(message.body)