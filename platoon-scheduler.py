import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Läs in de två Excel-filerna
schedule = 'schedule.xlsx'  # Byt till korrekt sökväg
answers = 'answers.xlsx'  # Byt till korrekt sökväg

# Öppna filen och använd det aktiva arket
wb1 = load_workbook(schedule)
ws1 = wb1.active
wb2 = load_workbook(answers)
ws2 = wb2.active

# Konvertera answers-ark till en pandas DataFrame
dfanswers = pd.DataFrame(ws2.values)

# Extrahera e-post, namn och telefonnummer från DataFrame-answers och lägg in i en array person_data
person_data = []
for index, row in dfanswers.iterrows():
    if index > 0 and pd.notna(row[1]) and pd.notna(row[2]) and pd.notna(row[3]):  # Hoppa över rubriker och kontrollera att cellerna inte är tomma
        person_data.append({
            "email": row[1],
            "name": row[2],
            "phone": row[3]
        })

# Skriv ut för att kontrollera att data har extraherats korrekt
print("Extraherad data:")
print(person_data)
print(f"Totalt antal personer: {len(person_data)}")

# Iterera genom cellerna i första arket och fyll i de gröna cellerna
# Detta program är optimerat för att fylla de tidigaste tiderna först i varje rum
person_index = 0
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
    for cell in row:
        # Kollar om cellen är färgad
        if cell.fill.fgColor.type == 'theme':
            # Kollar om det finns personer kvar
            if person_index < len(person_data):
                person = person_data[person_index]
                print(f"Fyller i cell {cell.coordinate} med data: {person}")
                cell.value = f"{person['email']}\n{person['name']}\n{person['phone']}"
                person_index += 1

# Kontrollera hur många som finns kvar och vilka som inte skrevs ut
if person_index != len(person_data):
    print(f"Det finns {len(person_data)-person_index} personer kvar")
    for person in person_data[person_index:]:
        print(person)

# Spara den uppdaterade arbetsboken till en ny fil
updated_file_path = 'updated_schedule.xlsx'  # Byt till önskad sökväg
wb1.save(updated_file_path)
